from flask import Flask, request, session, redirect, render_template_string
import requests
import os
from dotenv import load_dotenv
import bcrypt
import secrets 
from datetime import datetime, timedelta, date
import smtplib
from email.mime.text import MIMEText
from openpyxl import Workbook
from flask import send_file
from io import BytesIO

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev_key")

from datetime import datetime, timedelta
from flask import session, redirect, url_for, request

# durata massima sessione
app.permanent_session_lifetime = timedelta(minutes=30)

@app.before_request
def check_session_timeout():
    # escludi login e statici
    if request.endpoint in ("login", "static"):
        return

    if "user" in session:
        now = datetime.utcnow()
        last_activity = session.get("last_activity")

        if last_activity:
            elapsed = now - datetime.fromisoformat(last_activity)

            if elapsed.total_seconds() > 1800:
                session.clear()
                return redirect(url_for("login"))

        session["last_activity"] = now.isoformat()

SUPABASE_URL = "https://dlhayirunoremlpkyxlo.supabase.co"
SUPABASE_KEY = "sb_publishable_DZ69ih5L9IqvJmt44VUK4w_8uelJ5xU"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

LOGIN_HTML = """
<!DOCTYPE html>
<html lang="it">
<head>
    <meta charset="UTF-8">
    <title>Ferie & Permessi Team104</title>
    <style>
        body {
            margin: 0;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            background: url('https://images.unsplash.com/photo-1497366754035-f200968a6e72') no-repeat center center/cover;
        }

        .login-card {
            backdrop-filter: blur(12px);
            background: rgba(0, 0, 0, 0.65);
            padding: 40px;
            border-radius: 18px;
            width: 360px;
            color: white;
            box-shadow: 0 0 30px rgba(0,0,0,0.6);
            text-align: center;
        }

        .login-card h2 {
            margin-bottom: 10px;
            font-size: 28px;
            letter-spacing: 1px;
            color: #38bdf8;
        }

        .logo {
            width: 100px;
            margin-bottom: 20px;
        }

        input {
            width: 100%;
            padding: 12px;
            margin-bottom: 18px;
            border: none;
            border-radius: 8px;
            outline: none;
            font-size: 14px;
        }

        input[type="text"],
        input[type="password"] {
            background: #f1f5f9;
        }

        button {
            width: 100%;
            padding: 12px;
            border: none;
            border-radius: 8px;
            background-color: #38bdf8;
            color: black;
            font-weight: bold;
            cursor: pointer;
            transition: 0.3s;
        }

        button:hover {
            background-color: #0ea5e9;
        }

        .links {
            margin-top: 15px;
            font-size: 14px;
            color: #94a3b8;
        }

        .links a {
            color: #38bdf8;
            text-decoration: none;
        }

        .footer {
            margin-top: 25px;
            font-size: 12px;
            color: #94a3b8;
        }

        .team {
            color: #38bdf8;
            font-weight: bold;
        }
    </style>
</head>
<body>

    <form method="post" class="login-card">
        <!-- Logo -->
        <img src="{{ url_for('static', filename='images/logo.jpeg') }}" alt="Logo" class="logo">
        
        <h2>Login Ferie&Permessi</h2>

        <!-- Username -->
        Username:
        <input name="username" required><br>

        <!-- Password -->
        Password:
        <input name="password" type="password" required><br>

        <!-- Submit Button -->
        <button type="submit">Login</button>

        <!-- Links for Register & Password Recovery -->
        <div class="links">
            <a href="/register">Registrati</a> | 
            <a href="/forgot">Password smarrita?</a>
        </div>

        <div class="footer">
            Powered by <span class="team">Team104</span>
        </div>
    </form>

</body>
</html>
"""

@app.route("/register", methods=["GET", "POST"])
def register():

    # ================= POST (SALVATAGGIO UTENTE) =================
    if request.method == "POST":

        username = request.form.get("username")
        email = request.form.get("email")
        password = request.form.get("password")
        sector = request.form.get("sector")
        first_name = request.form.get("first_name")
        last_name = request.form.get("last_name")

        # 🔒 controllo base
        if not username or not email or not password or not sector:
            return "Tutti i campi sono obbligatori"

        # 🔎 evita duplicati username
        check = requests.get(
            f"{SUPABASE_URL}/rest/v1/users?username=eq.{username}",
            headers=HEADERS
        )

        if check.json():
            return "Username già registrato"

        # 🔐 hash password
        hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

        data = {
            "username": username,
            "email": email,
            "password": hashed,
            "role": "worker",
            "sector": sector,
            "first_name": first_name,
            "last_name": last_name
        }

        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/users",
            headers=HEADERS,
            json=data
        )

        print("REGISTER STATUS:", res.status_code)
        print("REGISTER RESPONSE:", res.text)

        if res.status_code not in [200, 201]:
            return "Errore registrazione"

        return """
        <h3>Registrazione completata ✅</h3>
        <a href="/">Torna al login</a>
        """

    # ================= GET (MOSTRA FORM) =================

    res = requests.get(
        f"{SUPABASE_URL}/rest/v1/users_available_free?select=username,sector,first_name,last_name",
        headers=HEADERS
    )

    users = res.json()

    return render_template_string("""
<!DOCTYPE html>
<html lang="it">
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">

    <style>
        body {
            margin: 0;
            font-family: Arial, sans-serif;
            background: url('https://images.unsplash.com/photo-1497366754035-f200968a6e72') no-repeat center center/cover;
            background-size: cover;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }

        .card {
            background: rgba(15, 23, 42, 0.92);
            padding: 25px;
            border-radius: 14px;
            width: 90%;
            max-width: 420px;
            color: white;
            box-shadow: 0 10px 30px rgba(0,0,0,0.5);
        }

        h2 {
            text-align: center;
            color: #38bdf8;
            margin-bottom: 20px;
        }

        label {
            font-size: 14px;
            display: block;
            margin-top: 10px;
            margin-bottom: 5px;
            color: #cbd5e1;
        }

        input, select {
            width: 100%;
            padding: 10px;
            border-radius: 8px;
            border: none;
            margin-bottom: 10px;
            outline: none;
        }

        button {
            width: 100%;
            padding: 10px;
            background: #3b82f6;
            color: white;
            border: none;
            border-radius: 8px;
            font-weight: bold;
            margin-top: 10px;
            cursor: pointer;
        }

        button:hover {
            background: #2563eb;
        }

        .link {
            text-align: center;
            margin-top: 12px;
        }

        .link a {
            color: #38bdf8;
            text-decoration: none;
            font-size: 13px;
        }
    </style>
</head>

<body>

<div class="card">

    <h2>Registrazione</h2>

    <form method="post">

        <label>Username</label>
        <select name="username" id="username" onchange="fillSector()" required>
            <option value="">Seleziona username</option>
            {% for u in users %}
                <option value="{{ u.username }}"
                        data-sector="{{ u.sector }}"
                        data-first="{{ u.first_name }}"
                        data-last="{{ u.last_name }}">
                    {{ u.username }}
                </option>
            {% endfor %}
        </select>

        <input type="hidden" name="sector" id="sector">
        <input type="hidden" name="first_name" id="first_name">
        <input type="hidden" name="last_name" id="last_name">

        <label>Email</label>
        <input name="email" type="email" required>

        <label>Password</label>
        <input name="password" type="password" required>

        <button type="submit">Registrati</button>

    </form>

    <div class="link">
        <a href="/">← Torna al login</a>
    </div>

</div>

<script>
function fillSector() {

    let select = document.getElementById("username");
    let option = select.options[select.selectedIndex];

    document.getElementById("sector").value = option.dataset.sector || "";
    document.getElementById("first_name").value = option.dataset.first || "";
    document.getElementById("last_name").value = option.dataset.last || "";
}
</script>

</body>
</html>
""", users=users)
def send_email(to, link):

    msg = MIMEText(f"Clicca qui per reimpostare la password:\n{link}")
    msg["Subject"] = "Reset Password"
    msg["From"] = "noreply.team104@gmail.com"
    msg["To"] = to

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login("noreply.team104@gmail.com", "turf tlus ngor onwr")
        server.send_message(msg)

@app.route("/forgot", methods=["GET", "POST"])
def forgot():

    if request.method == "POST":

        email = request.form["email"]

        # 🔎 controllo che email esista
        check = requests.get(
            f"{SUPABASE_URL}/rest/v1/users?email=eq.{email}",
            headers=HEADERS
        )

        user_data = check.json()

        if not user_data:
            return "Email non registrata"

        # 🔐 genera token
        token = secrets.token_urlsafe(32)
        expires = (datetime.utcnow() + timedelta(minutes=15)).isoformat()

        # 💾 salva token su users
        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/users?email=eq.{email}",
            headers=HEADERS,
            json={
                "reset_token": token
            }
        )

        print("RESET PATCH STATUS:", res.status_code)
        print("RESET PATCH RESPONSE:", res.text)

        # 🔗 link reset
        reset_link = f"https://attendance-app-9ozz.onrender.com/reset/{token}"

        send_email(email, reset_link)

        return """
        <h3>Ti abbiamo inviato una mail 📩</h3>
        <p>Controlla la tua casella e segui il link per reimpostare la password.</p>

        <a href="/" style="
            display:inline-block;
            margin-top:10px;
            padding:8px 12px;
            background:#3b82f6;
            color:white;
            text-decoration:none;
            border-radius:6px;
        ">
            Torna al login
        </a>
        """

    return """
<!DOCTYPE html>
<html lang="it">
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">

    <style>
        body {
            margin: 0;
            font-family: Arial, sans-serif;
            background: url('https://images.unsplash.com/photo-1497366754035-f200968a6e72') no-repeat center center/cover;
            background-size: cover;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }

        .card {
            background: rgba(15, 23, 42, 0.92);
            padding: 25px;
            border-radius: 14px;
            width: 90%;
            max-width: 420px;
            color: white;
            box-shadow: 0 10px 30px rgba(0,0,0,0.5);
        }

        h2 {
            text-align: center;
            color: #38bdf8;
            margin-bottom: 20px;
        }

        label {
            font-size: 14px;
            display: block;
            margin-top: 10px;
            margin-bottom: 5px;
            color: #cbd5e1;
        }

        input {
            width: 100%;
            padding: 10px;
            border-radius: 8px;
            border: none;
            margin-bottom: 10px;
            outline: none;
        }

        button {
            width: 100%;
            padding: 10px;
            background: #3b82f6;
            color: white;
            border: none;
            border-radius: 8px;
            font-weight: bold;
            margin-top: 10px;
            cursor: pointer;
        }

        button:hover {
            background: #2563eb;
        }

        .link {
            text-align: center;
            margin-top: 12px;
        }

        .link a {
            color: #38bdf8;
            text-decoration: none;
            font-size: 13px;
        }
    </style>
</head>

<body>

<div class="card">

    <h2>Password smarrita</h2>

    <form method="post">

        <label>Inserisci la tua email</label>
        <input name="email" type="email" required>

        <button type="submit">Invia</button>

    </form>

    <div class="link">
        <a href="/">← Torna al login</a>
    </div>

</div>

</body>
</html>
"""
@app.route("/reset/<token>", methods=["GET", "POST"])
def reset_password(token):

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json"
    }

    if request.method == "POST":

        new_password = request.form["password"].encode("utf-8")
        hashed = bcrypt.hashpw(new_password, bcrypt.gensalt()).decode("utf-8")

        # 🔎 1. CERCO UTENTE COL TOKEN
        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/users?reset_token=eq.{token}",
            headers=headers
        )

        print("RESET GET STATUS:", res.status_code)
        print("RESET GET TEXT:", res.text)

        if res.status_code != 200:
            return "Errore server durante verifica token"

        try:
            users = res.json()
        except:
            return "Errore lettura dati token"

        if not users:
            return "Token non valido o scaduto"

        user_id = users[0]["id"]

        # 🔄 2. UPDATE PASSWORD
        update = requests.patch(
            f"{SUPABASE_URL}/rest/v1/users?id=eq.{user_id}",
            headers=headers,
            json={
                "password": hashed,
                "reset_token": None
            }
        )

        print("UPDATE STATUS:", update.status_code)
        print("UPDATE TEXT:", update.text)

        if update.status_code not in [200, 204]:
            return f"Errore aggiornamento password: {update.text}"

        return """
<!DOCTYPE html>
<html lang="it">
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">

    <style>
        body {
            margin: 0;
            font-family: Arial, sans-serif;
            background: url('https://images.unsplash.com/photo-1497366754035-f200968a6e72') no-repeat center center/cover;
            background-size: cover;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }

        .card {
            background: rgba(15, 23, 42, 0.92);
            padding: 25px;
            border-radius: 14px;
            width: 90%;
            max-width: 420px;
            color: white;
            box-shadow: 0 10px 30px rgba(0,0,0,0.5);
            text-align: center;
        }

        h3 {
            color: #22c55e;
            font-size: 24px;
            margin-bottom: 20px;
        }

        a {
            display: inline-block;
            margin-top: 10px;
            padding: 8px 12px;
            background: #3b82f6;
            color: white;
            text-decoration: none;
            border-radius: 6px;
            font-weight: bold;
        }

        a:hover {
            background: #2563eb;
        }
    </style>
</head>

<body>

<div class="card">

    <h3>Password aggiornata con successo ✅</h3>

    <a href="/">Torna al login</a>

</div>

</body>
</html>
"""

    return """
<!DOCTYPE html>
<html lang="it">
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">

    <style>
        body {
            margin: 0;
            font-family: Arial, sans-serif;
            background: url('https://images.unsplash.com/photo-1497366754035-f200968a6e72') no-repeat center center/cover;
            background-size: cover;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }

        .card {
            background: rgba(15, 23, 42, 0.92);
            padding: 25px;
            border-radius: 14px;
            width: 90%;
            max-width: 420px;
            color: white;
            box-shadow: 0 10px 30px rgba(0,0,0,0.5);
            text-align: center;
        }

        h3 {
            color: #22c55e;
            font-size: 24px;
            margin-bottom: 20px;
        }

        input[type="password"] {
            width: 100%;
            padding: 10px;
            border-radius: 6px;
            border: 1px solid #ccc;
            margin-bottom: 20px;
            background-color: #1f2937;
            color: white;
        }

        button {
            padding: 10px 15px;
            background: #22c55e;
            color: white;
            border: none;
            border-radius: 6px;
            font-weight: bold;
            width: 100%;
            cursor: pointer;
        }

        button:hover {
            background: #16a34a;
        }
    </style>
</head>

<body>

<div class="card">

    <h3>Inserisci nuova password</h3>

    <form method="POST">
        <input type="password" name="password" required placeholder="Nuova password">
        <button type="submit">Reset Password</button>
    </form>

</div>

</body>
</html>
"""
# ---------------- LOGIN ----------------
@app.route("/", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/users?username=eq.{username}",
            headers=HEADERS
        )

        try:
            user_data = res.json()
        except:
            return "Errore server"

        if not user_data:
            return """
            <!DOCTYPE html>
            <html lang="it">
            <head>
            <meta charset="UTF-8">
            <title>Utente non trovato</title>
            <style>
                body {
                    margin: 0;
                    height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    background: linear-gradient(135deg, #1e3c72, #2a5298);
                    font-family: Arial, Helvetica, sans-serif;
                }
            
                .alert-box {
                    background: white;
                    padding: 40px 50px;
                    border-radius: 12px;
                    box-shadow: 0 15px 35px rgba(0,0,0,0.2);
                    text-align: center;
                    max-width: 420px;
                    animation: fadeIn 0.4s ease-in-out;
                }
            
                .alert-icon {
                    font-size: 50px;
                    color: #f39c12;
                    margin-bottom: 15px;
                }
            
                .alert-title {
                    font-size: 22px;
                    font-weight: bold;
                    margin-bottom: 10px;
                    color: #333;
                }
            
                .alert-message {
                    font-size: 16px;
                    color: #666;
                    margin-bottom: 25px;
                }
            
                .btn-back {
                    display: inline-block;
                    padding: 12px 25px;
                    background-color: #2a5298;
                    color: white;
                    text-decoration: none;
                    border-radius: 8px;
                    transition: 0.2s;
                    font-weight: bold;
                }
            
                .btn-back:hover {
                    background-color: #1e3c72;
                    transform: scale(1.05);
                }
            
                @keyframes fadeIn {
                    from { opacity: 0; transform: translateY(10px); }
                    to { opacity: 1; transform: translateY(0); }
                }
            </style>
            </head>
            <body>
            
            <div class="alert-box">
                <div class="alert-icon">⚠️</div>
                <div class="alert-title">Utente non trovato</div>
                <div class="alert-message">
                    Non esiste alcun account con questo username.<br>
                    Verifica di averlo digitato correttamente.
                </div>
                <a href="/" class="btn-back">Torna al login</a>
            </div>
            
            </body>
            </html>
            """

        db_user = user_data[0]

        if bcrypt.checkpw(password.encode(), db_user["password"].encode()):
            session.permanent = True
            session["user"] = db_user
            session["last_activity"] = datetime.utcnow().isoformat()
            role = db_user.get("role", "worker")

            session["user"] = db_user
            session["view"] = "manager" if role == "manager" else "worker"   
            return redirect("/dashboard")

        return """
        <!DOCTYPE html>
        <html lang="it">
        <head>
        <meta charset="UTF-8">
        <title>Login errore</title>
        <style>
            body {
                margin: 0;
                height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
                background: linear-gradient(135deg, #1e3c72, #2a5298);
                font-family: Arial, Helvetica, sans-serif;
            }
        
            .alert-box {
                background: white;
                padding: 40px 50px;
                border-radius: 12px;
                box-shadow: 0 15px 35px rgba(0,0,0,0.2);
                text-align: center;
                max-width: 420px;
                animation: fadeIn 0.4s ease-in-out;
            }
        
            .alert-icon {
                font-size: 50px;
                color: #e74c3c;
                margin-bottom: 15px;
            }
        
            .alert-title {
                font-size: 22px;
                font-weight: bold;
                margin-bottom: 10px;
                color: #333;
            }
        
            .alert-message {
                font-size: 16px;
                color: #666;
                margin-bottom: 25px;
            }
        
            .btn-back {
                display: inline-block;
                padding: 12px 25px;
                background-color: #2a5298;
                color: white;
                text-decoration: none;
                border-radius: 8px;
                transition: 0.2s;
                font-weight: bold;
            }
        
            .btn-back:hover {
                background-color: #1e3c72;
                transform: scale(1.05);
            }
        
            @keyframes fadeIn {
                from { opacity: 0; transform: translateY(10px); }
                to { opacity: 1; transform: translateY(0); }
            }
        </style>
        </head>
        <body>
        
        <div class="alert-box">
            <div class="alert-icon">⛔</div>
            <div class="alert-title">Credenziali non valide</div>
            <div class="alert-message">
                Lo username o la password inseriti non sono corretti.<br>
                Riprova.
            </div>
            <a href="/" class="btn-back">Torna al login</a>
        </div>
        
        </body>
        </html>
        """

    return render_template_string(LOGIN_HTML)


# ---------------- DASHBOARD ----------------
@app.route("/switch_view/<view>")
def switch_view(view):

    print("SWITCH CHIAMATO:", view)
    print("USER PRIMA:", session.get("user"))
    print("VIEW PRIMA:", session.get("view"))

    session["view"] = view

    print("VIEW DOPO:", session.get("view"))

    return redirect("/dashboard")

@app.route("/dashboard")
def dashboard():
    print(session)
    if "user" not in session:
        return redirect("/")
    view = session.get("view", "worker")
    user = session["user"]

    # ===== SWITCH SEMPRE VISIBILE AL MANAGER =====
    switch_html = ""
    
    if user.get("role") == "manager":
        current_view = session.get("view", "manager")
    
        switch_html = f"""
        <div style="margin-bottom:15px; padding:10px; background:#0f172a; border-radius:8px; display:flex; gap:10px; align-items:center;">
            <b style="color:white;">Vista Attuale:</b>
    
            <a href="/switch_view/manager" style="text-decoration:none;">
                <button style="cursor:pointer; padding:6px 12px; background:{'#22c55e' if current_view=='manager' else '#334155'}; color:white; border:none; border-radius:6px; font-weight:bold;">
                    👔 Manager
                </button>
            </a>
    
            <a href="/switch_view/worker" style="text-decoration:none;">
                <button style="cursor:pointer; padding:6px 12px; background:{'#3b82f6' if current_view=='worker' else '#334155'}; color:white; border:none; border-radius:6px; font-weight:bold;">
                    👷 Lavoratore
                </button>
            </a>
        </div>
        """

    sector = user["sector"]

    # ===== Recupero TUTTE le richieste pending =====
    res = requests.get(
        f"{SUPABASE_URL}/rest/v1/absences?status=eq.pending",
        headers=HEADERS
    )
    
    all_pending = res.json()
    
    # Conta quante richieste pending per ogni sector
    pending_by_sector = {}
    
    for req in all_pending:
        s = req["sector"]
        pending_by_sector[s] = pending_by_sector.get(s, 0) + 1

    # ================= CAPO =================
    

    if view == "manager":

        import json

        sector = request.args.get("sector")
        if not sector:
            sector = "all"
    
        params = {
            "select": "*"
        }
    
        if sector and sector != "all":
            params["sector"] = f"eq.{sector}"
    
        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/absences",
            headers=HEADERS,
            params=params
        )
    
        try:
            data = res.json()
        except:
            data = []

        # ================= PREPARAZIONE EVENTI CALENDARIO =================
        events = []

        for d in data:

            # FullCalendar usa END ESCLUSIVO → aggiungiamo 1 giorno per ferie
            if d.get("type") == "ferie":
                end_date = datetime.strptime(d["date_to"], "%Y-%m-%d") + timedelta(days=1)
                end_date = end_date.strftime("%Y-%m-%d")
            else:
                end_date = d["date_from"]

            events.append({
                "id": d["id"],
                "title": f"{d['worker_name']} - {d['type'].capitalize()}",
                "start": d["date_from"],
                "end": end_date,
                "color":
                    "#f59e0b" if d["status"] == "pending"
                    else "#22c55e" if d["status"] == "approved"
                    else "#ef4444",
                "extendedProps": {
                    "worker": d["worker_name"],
                    "type": d["type"],
                    "date_from": d["date_from"],
                    "date_to": d.get("date_to"),
                    "start_time": d.get("start_time"),
                    "end_time": d.get("end_time"),
                    "status": d["status"]
                }
            })

        events_json = json.dumps(events)
        # richieste pending SOLO del sector visualizzato
        pending_requests = [r for r in all_pending if r["sector"] == sector] if sector else []
        if user["role"] == "manager":

            # Recuperiamo il ruolo reale e la visualizzazione attuale
            user_role = user.get("role")
            current_view = session.get("view")
            
            # Se è la prima volta, impostiamo la vista predefinita in base al ruolo
            if not current_view:
                current_view = user_role
                session["view"] = current_view
            
            html = ""
            
            # Mostriamo lo switch SOLO se l'utente è un Manager (sia in modalità manager che worker)
            if user_role == "manager":
                html += f"""
                <div style="margin-bottom:15px; padding:10px; background:#0f172a; border-radius:8px; display:flex; gap:10px; align-items:center;">
                    <b style="color:white;">Vista Attuale:</b>
            
                    <a href="/switch_view/manager" style="text-decoration:none;">
                        <button style="cursor:pointer; padding:6px 12px; background:{'#22c55e' if current_view=='manager' else '#334155'}; color:white; border:none; border-radius:6px; font-weight:bold;">
                            👔 Manager
                        </button>
                    </a>
            
                    <a href="/switch_view/worker" style="text-decoration:none;">
                        <button style="cursor:pointer; padding:6px 12px; background:{'#3b82f6' if current_view=='worker' else '#334155'}; color:white; border:none; border-radius:6px; font-weight:bold;">
                            👷 Lavoratore
                        </button>
                    </a>
                </div>       

        <style>
        .sector-btn {{
            background:#2d89ef;
            color:white;
            padding:8px 14px;
            border:none;
            border-radius:6px;
            font-weight:bold;
            cursor:pointer;
        }}
        
        .alert-btn {{
            background:#e74c3c !important;
            animation:pulse 1.2s infinite;
        }}
        
        @keyframes pulse {{
            0% {{ box-shadow:0 0 0 0 rgba(231,76,60,0.7); }}
            70% {{ box-shadow:0 0 0 10px rgba(231,76,60,0); }}
            100% {{ box-shadow:0 0 0 0 rgba(231,76,60,0); }}
        }}
        
        .notification-box {{
            background:#1f2937;
            color:white;
            padding:12px;
            border-radius:8px;
            margin-bottom:15px;
        }}
        .selected-btn {{
            background:#0ea5e9 !important;
            border:2px solid white;
            transform:scale(1.05);
        }}
        </style>
        
        <div style="margin-bottom:15px; display:flex; gap:8px; flex-wrap:wrap;">
            <a href="/dashboard?sector=all">
                <button class="sector-btn
                {'selected-btn' if sector=='all' else ''}">
                Tutti
                </button>
            </a>
        
            <a href="/dashboard?sector=Dogane">
                <button class="sector-btn 
                {'alert-btn' if pending_by_sector.get('Dogane',0) > 0 else ''}
                {'selected-btn' if sector=='Dogane' else ''}">
                Dogane {'🔔 ' + str(pending_by_sector.get('Dogane',0)) if pending_by_sector.get('Dogane',0) > 0 else ''}
                </button>
            </a>
        
            <a href="/dashboard?sector=Syllabus">
                <button class="sector-btn 
                {'alert-btn' if pending_by_sector.get('Syllabus',0) > 0 else ''}
                {'selected-btn' if sector=='Syllabus' else ''}">
                Syllabus {'🔔 ' + str(pending_by_sector.get('Syllabus',0)) if pending_by_sector.get('Syllabus',0) > 0 else ''}
                </button>
            </a>
        
            <a href="/dashboard?sector=Unica">
                <button class="sector-btn 
                {'alert-btn' if pending_by_sector.get('Unica',0) > 0 else ''}
                {'selected-btn' if sector=='Unica' else ''}">
                Unica {'🔔 ' + str(pending_by_sector.get('Unica',0)) if pending_by_sector.get('Unica',0) > 0 else ''}
                </button>
            </a>
        
            <a href="/dashboard?sector=Accise">
                <button class="sector-btn 
                {'alert-btn' if pending_by_sector.get('Accise',0) > 0 else ''}
                {'selected-btn' if sector=='Accise' else ''}">
                Accise {'🔔 ' + str(pending_by_sector.get('Accise',0)) if pending_by_sector.get('Accise',0) > 0 else ''}
                </button>
            </a>
        
            <a href="/dashboard?sector=Fabbisogni">
                <button class="sector-btn 
                {'alert-btn' if pending_by_sector.get('Fabbisogni',0) > 0 else ''}
                {'selected-btn' if sector=='Fabbisogni' else ''}">
                Fabbisogni {'🔔 ' + str(pending_by_sector.get('Fabbisogni',0)) if pending_by_sector.get('Fabbisogni',0) > 0 else ''}
                </button>
            </a>
        
            <a href="/dashboard?sector=Bonus">
                <button class="sector-btn 
                {'alert-btn' if pending_by_sector.get('Bonus',0) > 0 else ''}
                {'selected-btn' if sector=='Bonus' else ''}">
                Bonus {'🔔 ' + str(pending_by_sector.get('Bonus',0)) if pending_by_sector.get('Bonus',0) > 0 else ''}
                </button>
            </a>
        
        
        
        <a href="/logout" style="
            display:inline-block;
            padding:8px 14px;
            background:#ef4444;
            color:white;
            text-decoration:none;
            border-radius:8px;
            font-weight:bold;
            font-family:Arial;
            transition:0.2s;
        ">
            Logout
        </a>
        <hr>
        <a href="/settings">
            ⚙️ Impostazioni account
        </a>
        <form method="GET" action="/export_excel" style="margin-bottom:20px; display:flex; gap:10px; align-items:end;">
            <div>
                <label style="color:white;">Dal:</label><br>
                <input type="date" name="date_from" required>
            </div>
            <div>
                <label style="color:white;">Al:</label><br>
                <input type="date" name="date_to" required>
            </div>
            <button type="submit" style="
                padding:8px 14px;
                background:#16a34a;
                color:white;
                border:none;
                border-radius:6px;
                font-weight:bold;
                cursor:pointer;">
                📥 Scarica Excel
            </button>
        </form>
        <!-- ================= FULLCALENDAR ================= -->
<link href='https://cdn.jsdelivr.net/npm/fullcalendar@6.1.11/index.global.min.css' rel='stylesheet' />
<script src='https://cdn.jsdelivr.net/npm/fullcalendar@6.1.11/index.global.min.js'></script>

<div id='calendar'></div>

<div id="eventModal" style="
    display:none;
    position:fixed;
    top:0; left:0;
    width:100%; height:100%;
    background:rgba(0,0,0,0.6);
    justify-content:center;
    align-items:center;
    z-index:9999;
">
    <div style="
        background:white;
        padding:20px;
        border-radius:10px;
        width:350px;
        font-family:Arial;
        position:relative;
    ">
        <span onclick="closeModal()" style="
            position:absolute;
            top:10px;
            right:15px;
            cursor:pointer;
            font-weight:bold;
        ">✖</span>
        <div id="modalBody"></div>
    </div>
</div>

<script>
document.addEventListener('DOMContentLoaded', function() {{

    var calendarEl = document.getElementById('calendar');

    var calendar = new FullCalendar.Calendar(calendarEl, {{
        initialView: "dayGridMonth",
        locale: "it",
        firstDay: 1,

        weekends: true,

        headerToolbar: {{
            left: 'prev,next today',
            center: 'title',
            right: 'timeGridDay,timeGridWeek'
        }},

        dayCellDidMount: function(info) {{
            const day = info.date.getDay();
            if (day === 0 || day === 6) {{
                info.el.style.backgroundColor = "#111827";
                info.el.style.opacity = "0.6";
            }}
        }},

        locale: 'it',
        slotMinTime: "08:00:00",
        slotMaxTime: "19:00:00",

        events: {events_json},

        eventClick: function(info) {{

            let e = info.event;

            let html = `
                <h3>${{e.extendedProps.worker}}</h3>
                <p><b>Tipo:</b> ${{e.extendedProps.type}}</p>
                <p><b>Data:</b> ${{e.extendedProps.date_from}} ${{e.extendedProps.date_to ? '→ ' + e.extendedProps.date_to : ''}}</p>
                <p><b>Orario:</b> ${{e.extendedProps.start_time ?? '09:00'}} - ${{e.extendedProps.end_time ?? '18:00'}}</p>
                <p><b>Stato:</b> ${{e.extendedProps.status}}</p>
                <br>
                <button onclick="handleAction('/approve/${{e.id}}')" style="padding:8px 12px;background:#22c55e;color:white;border:none;border-radius:6px;margin-right:8px;">✔ Approva</button>
                <button onclick="handleAction('/reject/${{e.id}}')" style="padding:8px 12px;background:#ef4444;color:white;border:none;border-radius:6px;">✖ Rifiuta</button>
            `;

            document.getElementById("modalBody").innerHTML = html;
            document.getElementById("eventModal").style.display = "flex";
        }},

        height: "auto"
    }});

    calendar.render();
}});
</script>

<script>
function closeModal() {{
    document.getElementById("eventModal").style.display = "none";
}}

function handleAction(url) {{
    fetch(url).then(() => {{
        closeModal();
        location.reload();
    }});
}}
</script>

<hr>
<h3 style="color:white">Lista richieste</h3>
        """

        # ================= LA TUA LISTA ORIGINALE =================    
        
        today = date.today()
        
        for d in data:
            start = datetime.strptime(d["date_from"], "%Y-%m-%d").date()
            end = datetime.strptime(d.get("date_to", d["date_from"]), "%Y-%m-%d").date()
        
            # Nascondi solo richieste completamente passate
            if end < today:
                continue
        
            color = "#f59e0b" if d["status"] == "pending" else "#22c55e" if d["status"] == "approved" else "#ef4444"
        
            if d.get("type") == "ferie":
                date_display = f'{d.get("date_from")} → {d.get("date_to")}'
                time_display = "09:00 - 18:00"
            else:
                date_display = d.get("date_from")
                time_display = f'{d.get("start_time")} - {d.get("end_time")}'
        
            html += f"""
            <div style="
                background:#0f172a;
                padding:12px;
                border-radius:10px;
                margin-bottom:10px;
                color:white;
                box-shadow:0 4px 12px rgba(0,0,0,0.4);
                text-align:left;
                width:100%;
            ">
                <b>{d["worker_name"]}</b><br>
                📅 {date_display}<br>
                🏷 {d.get("type","")}<br>
                ⏰ {time_display}<br>
                Stato: <span style="color:{color}">{d["status"]}</span><br><br>
        
                <a href="/approve/{d["id"]}" style="color:#22c55e">✔ Approva</a> |
                <a href="/reject/{d["id"]}" style="color:#ef4444">✖ Rifiuta</a>
            </div>
            """
        
        return html

    # ================= LAVORATORE =================
    else:

        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/absences?worker_name=eq.{user['username']}",
            headers=HEADERS
        )
        data = res.json()

        html = f"""
        <h2 style="color:#38bdf8">Benvenuto {user['first_name']}</h2>
        <a href="/logout" style="
            display:inline-block;
            padding:8px 14px;
            background:#ef4444;
            color:white;
            text-decoration:none;
            border-radius:8px;
            font-weight:bold;
            font-family:Arial;
            transition:0.2s;
        ">
            Logout
        </a>
        <a href="/settings">
            ⚙️ Impostazioni account
        </a>
        <hr>

        <h3>➕ Inserisci assenza</h3>

        <form method="post" action="/add_absence" style="
            background:#111827;
            padding:15px;
            border-radius:10px;
            color:white;
        ">

          Tipo:
          <select name="type" id="type" onchange="toggleAddForm()">
            <option value="ferie">Ferie</option>
            <option value="permesso">Permesso</option>
          </select><br><br>

          <div id="singleDate">
            Data: <input type="date" name="date"><br><br>
        </div>
        
        <div id="rangeDate" style="display:none;">
            Dal: <input type="date" name="date_from"><br><br>
            Al: <input type="date" name="date_to"><br><br>
        </div>

          Dalle: <input type="time" name="start_time" id="start" min="09:00" max="18:00"><br><br>
          
          Alle: <input type="time" name="end_time" id="end" min="09:00" max="18:00"><br><br>

        <button id="submitBtn" type="submit" disabled
        style="background:#3b82f6;color:white;padding:6px;border:none;border-radius:6px;opacity:0.5;">
        Invia
        </button>
        </form>

        <script>
        function toggleAddForm(){{

            let type = document.getElementById("type").value;
        
            let start = document.getElementById("start");
            let end = document.getElementById("end");
        
            let singleDate = document.getElementById("singleDate");
            let rangeDate = document.getElementById("rangeDate");
        
            if(type === "ferie"){{
                start.disabled = true;
                end.disabled = true;
                start.value = "";
                end.value = "";
        
                singleDate.style.display = "none";
                rangeDate.style.display = "block";
        
            }} else {{
                start.disabled = false;
                end.disabled = false;
        
                singleDate.style.display = "block";
                rangeDate.style.display = "none";
            }}
        
            validateForm();
        }}

        function blockWeekendDates() {{
    document.querySelectorAll("input[type='date']").forEach(input => {{
        input.addEventListener("input", function () {{
            const day = new Date(this.value).getDay();
            if (day === 0 || day === 6) {{
                alert("Weekend non selezionabile");
                this.value = "";
            }}
        }});
    }});
}}

window.addEventListener("DOMContentLoaded", blockWeekendDates);

        function validateForm(){{

            let type = document.getElementById("type").value;
        
            let start = document.getElementById("start").value;
            let end = document.getElementById("end").value;
        
            let submitBtn = document.getElementById("submitBtn");
        
            let valid = true;
        
            if(type === "ferie"){{
                let from = document.querySelector("input[name='date_from']").value;
                let to = document.querySelector("input[name='date_to']").value;
        
                if(!from || !to) valid = false;
        
            }} else {{
                let date = document.querySelector("input[name='date']").value;
        
                if(!date || !start || !end) valid = false;
            }}
        
            submitBtn.disabled = !valid;
            submitBtn.style.opacity = valid ? "1" : "0.5";
        }}

            document.querySelectorAll("input, select").forEach(el => {{
        el.addEventListener("input", validateForm);
    }});
            window.addEventListener("DOMContentLoaded", function () {{
            toggleAddForm();
        }});
        </script>

        <hr>
        <h3 style="color:#38bdf8;margin-bottom:10px;">📅 Calendario assenze</h3>
        
        </div>
        <link href='https://cdn.jsdelivr.net/npm/fullcalendar@6.1.11/index.global.min.css' rel='stylesheet' />
        <script src='https://cdn.jsdelivr.net/npm/fullcalendar@6.1.11/index.global.min.js'></script>
        
        <style>
            html, body {{
                margin: 0;
                padding: 0;
                height: 100%;
                font-family: Arial, Helvetica, sans-serif;
                background: #f4f6f9;
            }}
        
            .page-wrapper {{
                padding: 30px;
                height: 100%;
                box-sizing: border-box;
            }}
        
            .calendar-card {{
                background: white;
                border-radius: 14px;
                box-shadow: 0 15px 35px rgba(0,0,0,0.12);
                padding: 20px;
                height: calc(100% - 60px);
            }}
        
            #calendar {{
                height: 100%;
            }}
        </style>
        
        <div class="page-wrapper">
            <div class="calendar-card">
                <div id="calendar"></div>
            </div>
        </div>
        """

        for d in data:

            status = d.get("status", "pending")
            color = "#f59e0b" if status == "pending" else "#22c55e" if status == "approved" else "#ef4444"
        
            html += f"""
            <div class="card" style="
                background:linear-gradient(135deg,#1e293b,#0f172a);
                padding:12px;
                border-radius:10px;
                margin-bottom:10px;
                color:white;
                box-shadow:0 6px 15px rgba(0,0,0,0.3)
            ">
        
                <input type="hidden" class="id" value='{d["id"]}'>
        
                <b>Stato:
                    <span style="color:{color}; font-weight:bold;">
                        {status.upper()}
                    </span>
                </b><br><br>
        
                Tipo:
                <select class="type">
                    <option value="ferie" {"selected" if d.get("type")=="ferie" else ""}>Ferie</option>
                    <option value="permesso" {"selected" if d.get("type")=="permesso" else ""}>Permesso</option>
                </select><br>
        
                Data:<br>
                {f"""
                Dal: <input type='date' class='date_from' value='{d.get("date_from","")}'><br>
                Al: <input type='date' class='date_to' value='{d.get("date_to","")}'><br>
                """ if d.get("type")=="ferie" else f"""
                <input type='date' class='date' value='{d.get("date_from","")}'><br>
                """}
        
                Dalle:
                <input type="time" class="start" value='{d.get("start_time","")}'><br>
        
                Alle:
                <input type="time" class="end" value='{d.get("end_time","")}'><br><br>
        
                <button onclick="update(this)" style="background:#3b82f6;color:white;">Modifica</button>
                <button onclick="remove(this)" style="background:#ef4444;color:white;">Elimina</button>
            </div>
            """

        html += """
<script>


function remove(btn){
    let card = btn.closest(".card");
    let id = card.querySelector(".id").value;

    fetch("/delete/" + id)
        .then(res => res.text())
        .then(() => {
            card.remove(); // sparisce subito senza reload
        });
}


function toggleRow(card){

    let type = card.querySelector(".type").value;
    let start = card.querySelector(".start");
    let end = card.querySelector(".end");

    if(type === "ferie"){
        start.disabled = true;
        end.disabled = true;
        start.value = "";
        end.value = "";
    } else {
        start.disabled = false;
        end.disabled = false;
    }
}

window.addEventListener("DOMContentLoaded", function() {

    document.querySelectorAll(".card").forEach(function(c){
        toggleRow(c);

        c.querySelector(".type").addEventListener("change", function(){
            toggleRow(c);
        });
    });

});

function update(btn){

    let card = btn.closest(".card");
    let type = card.querySelector(".type").value;

    let payload = {
        id: card.querySelector(".id").value,
        type: type,
        start_time: card.querySelector(".start")?.value || null,
        end_time: card.querySelector(".end")?.value || null,
        status: "pending"
    };

    if (type === "ferie") {

        let fromEl = card.querySelector(".date_from");
        let toEl = card.querySelector(".date_to");

        payload.date_from = fromEl ? fromEl.value : null;
        payload.date_to = toEl ? toEl.value : null;

    } else {

        let dateEl = card.querySelector(".date");

        payload.date_from = dateEl ? dateEl.value : null;
        payload.date_to = dateEl ? dateEl.value : null;
    }

    fetch("/update_absence", {
        method: "POST",
        headers: {"Content-Type": "application/json"},
        body: JSON.stringify(payload)
    })
    .then(res => res.json())
    .then(() => location.reload());
}

function remove(btn){
    let id = btn.parentElement.querySelector(".id").value;

    fetch("/delete/" + id)
    .then(() => location.reload());
}

const data = {{ data | tojson }};

let currentDate = new Date();

document.addEventListener("DOMContentLoaded", function () {

    if (!window.FullCalendar) {
        console.error("FullCalendar non caricato");
        return;
    }

    const calendarEl = document.getElementById("calendar");

    const calendar = new FullCalendar.Calendar(calendarEl, {
        initialView: "dayGridMonth",
        locale: "it",
        firstDay: 1,

        headerToolbar: {
            left: 'prev,next today',
            center: 'title',
            right: 'dayGridMonth,timeGridWeek,timeGridDay'
        },
        

        weekends: true,

        dayCellDidMount: function(info) {
            const day = info.date.getDay();
            if (day === 0 || day === 6) {
                info.el.style.backgroundColor = "#0b1220";
                info.el.style.opacity = "0.5";
            }
        },

        events: (typeof data !== "undefined" ? data : []).map(d => {

            if (d.type === "ferie") {
                return {
                    title: "Ferie",
                    start: d.date_from,
                    end: new Date(new Date(d.date_to).getTime() + 86400000).toISOString().split("T")[0],
                    color: d.status === "approved" ? "#22c55e"
                          : d.status === "rejected" ? "#ef4444"
                          : "#f59e0b"
                };
            }

            return {
                title: "Permesso",
                start: d.date_from,
                color: d.status === "approved" ? "#22c55e"
                      : d.status === "rejected" ? "#ef4444"
                      : "#f59e0b"
            };
        })
    });

    calendar.render();
});

</script>
"""
        
        html = switch_html + html
        return render_template_string(html, data=data)


# ---------------- ADD ----------------
@app.route("/add_absence", methods=["POST"])
def add_absence():

    if "user" not in session:
        return redirect("/")

    user = session["user"]
    real_role = user.get("role")  # ← RUOLO VERO

    absence_type = request.form["type"]

    # ---------------- FERIE ----------------
    if absence_type == "ferie":
        date_from = request.form["date_from"]
        date_to = request.form["date_to"]
        start_time = None
        end_time = None

    # ---------------- PERMESSO ----------------
    else:
        date_from = request.form["date"]
        date_to = request.form["date"]
        start_time = request.form["start_time"]
        end_time = request.form["end_time"]

    status = "approved" if real_role == "manager" else "pending"

    data = {
        "worker_name": user["username"],
        "sector": user["sector"],
        "date_from": date_from,
        "date_to": date_to,
        "type": absence_type,
        "start_time": start_time,
        "end_time": end_time,
        "status": status
    }

    requests.post(
        f"{SUPABASE_URL}/rest/v1/absences",
        headers=HEADERS,
        json=data
    )

    return redirect("/dashboard")


@app.route("/update_absence", methods=["POST"])
def update_absence():

    if "user" not in session:
        return {"ok": False}, 401

    user = session["user"]
    real_role = user.get("role")  # ← RUOLO VERO
    data = request.json

    status = "approved" if real_role == "manager" else "pending"

    payload = {
        "type": data["type"],
        "start_time": data.get("start_time"),
        "end_time": data.get("end_time"),
        "status": status
    }

    # ---------------- FERIE ----------------
    if data["type"] == "ferie":
        payload["date_from"] = data.get("date_from")
        payload["date_to"] = data.get("date_to")

    # ---------------- PERMESSO ----------------
    else:
        payload["date_from"] = data.get("date_from")
        payload["date_to"] = data.get("date_to")

    requests.patch(
        f"{SUPABASE_URL}/rest/v1/absences?id=eq.{data['id']}",
        headers=HEADERS,
        json=payload
    )

    return {"ok": True}

# ---------------- DELETE ----------------

@app.route("/delete/<int:id>")
def delete_absence(id):

    print("DELETE CHIAMATA CON ID:", id)

    requests.delete(
        f"{SUPABASE_URL}/rest/v1/absences?id=eq.{id}",
        headers=HEADERS
    )

    return ("ok", 200)

# ---------------- APPROVE ----------------
@app.route("/approve/<id>")
def approve(id):

    r = requests.patch(
        f"{SUPABASE_URL}/rest/v1/absences?id=eq.{id}",
        headers=HEADERS,
        json={"status": "approved"}
    )

    print("APPROVE STATUS:", r.status_code)
    print("APPROVE RESPONSE:", r.text)

    return redirect("/dashboard")


# ---------------- REJECT ----------------
@app.route("/reject/<id>")
def reject(id):

    requests.patch(
        f"{SUPABASE_URL}/rest/v1/absences?id=eq.{id}",
        headers=HEADERS,
        json={"status": "rejected"}
    )

    return redirect("/dashboard")
    
# ---------------- DOWNLOAD EXCEL ----------------
@app.route("/export_excel")
def export_excel():

    if "user" not in session:
        return redirect("/")

    user = session["user"]

    # Solo manager può esportare
    if user.get("role") != "manager":
        return "Non autorizzato", 403

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    # Recupera assenze nel range date
    params = {
        "date_from": f"gte.{date_from}",
        "date_to": f"lte.{date_to}",
        "select": "*",
        "order": "sector"
    }

    res = requests.get(
        f"{SUPABASE_URL}/rest/v1/absences",
        headers=HEADERS,
        params=params
    )

    data = res.json()

    # ===== CREA EXCEL =====
    wb = Workbook()

    # Raggruppa per sector
    sectors = {}
    for r in data:
        s = r["sector"]
        if s not in sectors:
            sectors[s] = []
        sectors[s].append(r)

    # Crea un foglio per ogni sector
    for sector, records in sectors.items():

        ws = wb.create_sheet(title=sector)

        # Intestazioni
        ws.append([
            "Lavoratore",
            "Tipo",
            "Dal",
            "Al",
            "Ora Inizio",
            "Ora Fine",
            "Stato"
        ])

        # Righe
        for r in records:
            ws.append([
                r["worker_name"],
                r["type"],
                r["date_from"],
                r["date_to"],
                r.get("start_time"),
                r.get("end_time"),
                r["status"]
            ])

    # Rimuove il foglio vuoto iniziale creato da default
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)

    # Salva in memoria
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"report_assenze_{date_from}_{date_to}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------------- SETTINGS ----------------
@app.route("/settings", methods=["GET", "POST"])
def settings():
    if "user_id" not in session:
        return redirect("/login")

    user_id = session["user_id"]

    if request.method == "POST":
        new_email = request.form.get("email")
        new_password = request.form.get("password")

        update_data = {}

        if new_email:
            update_data["email"] = new_email

        if new_password:
            hashed = bcrypt.hashpw(new_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
            update_data["password"] = hashed
            if new_password != request.form.get("confirm"):
                return "Password non corrispondono"

        if update_data:
            url = f"{SUPABASE_URL}/rest/v1/users?id=eq.{user_id}"
            headers = {
                "apikey": SUPABASE_KEY,
                "Authorization": f"Bearer {SUPABASE_KEY}",
                "Content-Type": "application/json",
                "Prefer": "return=minimal"
            }

            requests.patch(url, json=update_data, headers=headers)

        return redirect("/dashboard")

    return render_template_string("""
    <h2>⚙️ Impostazioni account</h2>

    <form method="POST">
        <label>Email</label><br>
        <input type="email" name="email" placeholder="Nuova email"><br><br>

        <label>Nuova password</label><br>
        <input type="password" name="password" placeholder="Nuova password"><br><br>
        <input type="password" name="confirm" placeholder="Conferma password">

        <button type="submit">Salva modifiche</button>
    </form>

    <br>
    <a href="/dashboard">⬅ Torna indietro</a>
    """)

# ---------------- LOGOUT ----------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ---------------- KEEP SERVER AWAKE ----------------
@app.route("/health")
def health():
    return "OK", 200


# ---------------- RUN ----------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
