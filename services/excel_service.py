from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO


def generate_excel(data):

    # ===== CREA FILE =====
    wb = Workbook()

    # colori stato (opzionale ma PRO)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    header_font = Font(bold=True)

    # ===== RAGGRUPPA PER SETTORE =====
    sectors = {}
    for r in data:
        sector = r.get("sector", "Altro")
        sectors.setdefault(sector, []).append(r)

    # ===== CREA FOGLI =====
    for sector, records in sectors.items():

        ws = wb.create_sheet(title=sector)

        # intestazione
        headers = [
            "Lavoratore",
            "Tipo",
            "Dal",
            "Al",
            "Ora Inizio",
            "Ora Fine",
            "Stato"
        ]

        ws.append(headers)

        # stile intestazione
        for col in ws[1]:
            col.font = header_font

        # ===== DATI =====
        for r in records:

            row = [
                r.get("worker_name"),
                r.get("type"),
                r.get("date_from"),
                r.get("date_to"),
                r.get("start_time"),
                r.get("end_time"),
                r.get("status")
            ]

            ws.append(row)

            # colora stato
            last_row = ws.max_row
            status_cell = ws.cell(row=last_row, column=7)

            if r.get("status") == "approved":
                status_cell.fill = green_fill
            elif r.get("status") == "rejected":
                status_cell.fill = red_fill
            else:
                status_cell.fill = yellow_fill

        # auto width colonne
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2

    # rimuovi foglio default
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # ===== CREA STREAM =====
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return stream
